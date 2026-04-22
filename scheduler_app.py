"""
Internal use — AL-Rafidain Group
Iraqi Labor Law Workforce Scheduler
Ref: Law No. 37/2015
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import tkinter.font as tkfont
from dataclasses import dataclass, field, asdict
from datetime import datetime, time, date, timedelta
from typing import Dict, List, Optional, Any
import openpyxl
from openpyxl.cell.cell import MergedCell
import csv
import os

# --- DATAMODEL ---

@dataclass
class Employee:
    emp_id: str
    name: str
    role: str
    department: str
    contract_type: str = "Permanent"
    contracted_weekly_hrs: float = 48.0
    shift_eligibility: str = "All"
    is_hazardous: int = 0
    is_industrial_rotating: int = 0
    hour_exempt: int = 0
    fixed_rest_day: int = 6  # Friday default
    phone: str = ""
    hire_date: Optional[date] = None
    notes: str = ""

@dataclass
class Shift:
    code: str
    name: str
    start_time: time
    end_time: time
    duration_hrs: float
    break_min: int = 0
    is_industrial: int = 0
    is_hazardous: int = 0
    is_work: int = 1
    description: str = ""

@dataclass
class PublicHoliday:
    dt: date
    name: str
    h_type: str = "Fixed"
    legal_ref: str = "Official Holidays Law No. 11/2010 Art. 1"

@dataclass
class Config:
    company: str = "AL-Rafidain Group"
    year: int = 2026
    month: int = 1
    days_in_month: int = 31
    month_start_date: Optional[date] = None
    weekend_policy: str = "Friday Only"
    weekly_rest_day_primary: int = 6
    weekly_rest_day_secondary: Optional[int] = None
    continuous_shifts_mode: str = "OFF"
    coverage_min: int = 10
    max_consec_work_days: int = 6

@dataclass
class Violation:
    emp_id: str
    day: int
    rule: str
    article: str
    message: str

# --- COMPLIANCE ENGINE ---

class ComplianceEngine:
    @staticmethod
    def check_compliance(store: 'DataStore') -> List[Violation]:
        violations = []
        shifts_by_code = {s.code: s for s in store.shifts}
        holidays_dates = {h.dt for h in store.holidays}
        config = store.config
        
        for emp in store.employees:
            schedule = store.schedule.get(emp.emp_id, {})
            
            # --- Rule 1: Daily Hours (Art. 67) ---
            if not emp.hour_exempt:
                for day, code in schedule.items():
                    if code in shifts_by_code:
                        s = shifts_by_code[code]
                        cap = 6 if emp.is_hazardous or s.is_hazardous else 8
                        if s.duration_hrs > cap:
                            violations.append(Violation(
                                emp.emp_id, day, "Daily hours cap", "Art. 67",
                                f"Worked {s.duration_hrs}hrs (Cap: {cap}hrs per Art. 67)"
                            ))

            # Prepare for weekly/consecutive checks
            work_flags = [] # (day, worked_hrs, is_work_code)
            for d in range(1, config.days_in_month + 1):
                code = schedule.get(d, "")
                hrs = 0.0
                is_work = False
                if code in shifts_by_code:
                    s = shifts_by_code[code]
                    hrs = s.duration_hrs if s.is_work else 0.0
                    is_work = bool(s.is_work)
                work_flags.append((d, hrs, is_work))

            # --- Rule 2: Weekly Hours (Art. 70) ---
            # --- Rule 5: Weekly Rest Day (Art. 72) ---
            if not emp.hour_exempt:
                for i in range(len(work_flags)):
                    window = work_flags[i:i+7]
                    total_hrs = sum(x[1] for x in window)
                    has_rest = any(not x[2] for x in window)
                    
                    cap_weekly = 36 if emp.is_hazardous else 48
                    if total_hrs > cap_weekly:
                        violations.append(Violation(
                            emp.emp_id, window[0][0], "Weekly hours cap", "Art. 70",
                            f"Rolling week total {total_hrs}hrs exceeds {cap_weekly}hrs"
                        ))
                    
                    if len(window) == 7 and not has_rest:
                        violations.append(Violation(
                            emp.emp_id, window[6][0], "Weekly rest day", "Art. 72",
                            "No rest day in 7-day rolling window"
                        ))

            # --- Rule 4: Consecutive Work Days (Art. 71, 72) ---
            consec = 0
            for d_idx, hrs, is_work in work_flags:
                if is_work:
                    consec += 1
                else:
                    consec = 0
                if consec > config.max_consec_work_days:
                    violations.append(Violation(
                        emp.emp_id, d_idx, f"Consecutive work days", "Art. 71 §5",
                        f"Work exceeds {config.max_consec_work_days} consecutive days"
                    ))

            # --- Rule 6: Holiday OT (Art. 74) ---
            for d, hrs, is_work in work_flags:
                # Construct date
                try:
                    curr_date = date(config.year, config.month, d)
                    if is_work and curr_date in holidays_dates:
                        code = schedule.get(d, "")
                        if "OT" not in code and "PH" not in code:
                            violations.append(Violation(
                                emp.emp_id, d, "Holiday OT flag", "Art. 74",
                                "Worked on public holiday without OT marker"
                            ))
                except ValueError:
                    continue

        return violations

# --- DATA STORE ---

class DataStore:
    def __init__(self):
        self.filename: Optional[str] = None
        self.employees: List[Employee] = []
        self.shifts: List[Shift] = []
        self.holidays: List[PublicHoliday] = []
        self.config = Config()
        self.schedule: Dict[str, Dict[int, str]] = {}
        self.dirty = False

    def load(self, path: str):
        self.filename = path
        wb = openpyxl.load_workbook(path, data_only=True)
        
        # Load Config
        if "Config" in wb.sheetnames:
            sh = wb["Config"]
            cfg_data = {}
            for r in range(1, sh.max_row + 1):
                key = sh.cell(r, 1).value
                val = sh.cell(r, 2).value
                if key: cfg_data[key] = val
            
            self.config.company = cfg_data.get("Company", "AL-Rafidain Group")
            self.config.year = int(cfg_data.get("Scheduling Year", 2026))
            self.config.month = int(cfg_data.get("Scheduling Month", 1))
            self.config.days_in_month = int(cfg_data.get("Days In Month", 31))
            self.config.coverage_min = int(cfg_data.get("Coverage target", 10))
            self.config.max_consec_work_days = int(cfg_data.get("Max consecutive work days", 6))
            self.config.continuous_shifts_mode = cfg_data.get("Continuous Successive Shifts (Art. 71 §1(B))", "OFF")

        # Load Employees
        self.employees = []
        if "Employees" in wb.sheetnames:
            sh = wb["Employees"]
            footer = self._first_footer_row(sh)
            for r in range(2, footer):
                row = [sh.cell(r, c).value for c in range(1, 15)]
                if not row[0]: continue
                self.employees.append(Employee(
                    emp_id=str(row[0]), name=str(row[1]), role=str(row[2]),
                    department=str(row[3]), contract_type=str(row[4] or "Permanent"),
                    contracted_weekly_hrs=float(row[5] or 48.0),
                    shift_eligibility=str(row[6] or "All"),
                    is_hazardous=int(row[7] or 0), is_industrial_rotating=int(row[8] or 0),
                    hour_exempt=int(row[9] or 0), fixed_rest_day=int(row[10] or 6),
                    phone=str(row[11] or ""), hire_date=row[12], notes=str(row[13] or "")
                ))

        # Load Shifts
        self.shifts = []
        if "ShiftsConfig" in wb.sheetnames:
            sh = wb["ShiftsConfig"]
            footer = self._first_footer_row(sh)
            for r in range(2, footer):
                row = [sh.cell(r, c).value for c in range(1, 11)]
                if not row[0]: continue
                # Handle time parsing
                s_t = row[2] if isinstance(row[2], time) else time(8,0)
                e_t = row[3] if isinstance(row[3], time) else time(16,0)
                self.shifts.append(Shift(
                    code=str(row[0]), name=str(row[1]), start_time=s_t, end_time=e_t,
                    duration_hrs=float(row[4] or 8.0), break_min=int(row[5] or 0),
                    is_industrial=int(row[6] or 0), is_hazardous=int(row[7] or 0),
                    is_work=int(row[8] or 1), description=str(row[9] or "")
                ))

        # Load Holidays
        self.holidays = []
        if "PublicHolidays" in wb.sheetnames:
            sh = wb["PublicHolidays"]
            for r in range(2, sh.max_row + 1):
                dt = sh.cell(r, 1).value
                if not dt: continue
                self.holidays.append(PublicHoliday(
                    dt=dt if isinstance(dt, date) else date(2026, 1, 1),
                    name=str(sh.cell(r, 2).value or ""),
                    h_type=str(sh.cell(r, 3).value or "Fixed"),
                    legal_ref=str(sh.cell(r, 4).value or "")
                ))

        # Load Schedule
        self.schedule = {}
        if "Schedule" in wb.sheetnames:
            sh = wb["Schedule"]
            footer = self._first_footer_row(sh)
            for r in range(9, footer):
                emp_id = sh.cell(r, 1).value
                if not emp_id: continue
                emp_sched = {}
                for d in range(1, 32):
                    val = sh.cell(r, d + 5).value
                    if val:
                        emp_sched[d] = str(val)
                self.schedule[str(emp_id)] = emp_sched

        self.dirty = False

    def save(self, path: Optional[str] = None):
        target = path or self.filename
        if not target: return
        
        # Load or create
        if os.path.exists(target):
            wb = openpyxl.load_workbook(target)
        else:
            messagebox.showerror("Error", "Template file missing. Cannot save without original formulas.")
            return

        # Write Config
        sh = wb["Config"] if "Config" in wb.sheetnames else wb.create_sheet("Config")
        cfg_lines = [
            ("Company", self.config.company),
            ("Scheduling Year", self.config.year),
            ("Scheduling Month", self.config.month),
            ("Days In Month", self.config.days_in_month),
            ("Coverage target", self.config.coverage_min),
            ("Max consecutive work days", self.config.max_consec_work_days),
            ("Continuous Successive Shifts (Art. 71 §1(B))", self.config.continuous_shifts_mode)
        ]
        for i, (k, v) in enumerate(cfg_lines, 1):
            self._safe_set(sh, i, 1, k)
            self._safe_set(sh, i, 2, v)

        # Write Employees
        sh = wb["Employees"]
        footer = self._first_footer_row(sh)
        # Clear existing rows up to footer
        for r in range(2, footer):
            for c in range(1, 15): self._safe_set(sh, r, c, None)
        for i, emp in enumerate(self.employees[:footer-2], 2):
            vals = [emp.emp_id, emp.name, emp.role, emp.department, emp.contract_type,
                    emp.contracted_weekly_hrs, emp.shift_eligibility, emp.is_hazardous,
                    emp.is_industrial_rotating, emp.hour_exempt, emp.fixed_rest_day,
                    emp.phone, emp.hire_date, emp.notes]
            for c, v in enumerate(vals, 1): self._safe_set(sh, i, c, v)

        # Write Shifts
        sh = wb["ShiftsConfig"]
        footer = self._first_footer_row(sh)
        for r in range(2, footer):
            for c in range(1, 11): self._safe_set(sh, r, c, None)
        for i, s in enumerate(self.shifts[:footer-2], 2):
            vals = [s.code, s.name, s.start_time, s.end_time, s.duration_hrs,
                    s.break_min, s.is_industrial, s.is_hazardous, s.is_work, s.description]
            for c, v in enumerate(vals, 1): self._safe_set(sh, i, c, v)

        # Write Schedule (preserve formulas in A-E)
        sh = wb["Schedule"]
        footer = self._first_footer_row(sh)
        # Just write the day columns 6-36 for row 9 to footer-1
        for r in range(9, footer):
            # Find which employee this row matches positionally based on our loaded list
            # We assume rows 9+ correspond to self.employees[0], [1], etc.
            idx = r - 9
            if idx < len(self.employees):
                emp = self.employees[idx]
                sched = self.schedule.get(emp.emp_id, {})
                for d in range(1, 32):
                    self._safe_set(sh, r, d + 5, sched.get(d, None))
            else:
                # Clear extra rows
                for d in range(1, 32): self._safe_set(sh, r, d + 5, None)

        # Write Holidays
        sh = wb["PublicHolidays"]
        # Clear
        for r in range(2, sh.max_row+1):
            for c in range(1, 5): self._safe_set(sh, r, c, None)
        for i, h in enumerate(self.holidays, 2):
            self._safe_set(sh, i, 1, h.dt)
            self._safe_set(sh, i, 2, h.name)
            self._safe_set(sh, i, 3, h.h_type)
            self._safe_set(sh, i, 4, h.legal_ref)

        wb.save(target)
        self.filename = target
        self.dirty = False

    def _first_footer_row(self, sh):
        if not sh.merged_cells.ranges: return sh.max_row + 1
        return min(mr.min_row for mr in sh.merged_cells.ranges)

    def _safe_set(self, sh, r, c, value):
        cell = sh.cell(row=r, column=c)
        if isinstance(cell, MergedCell): return False
        try:
            cell.value = value
            return True
        except AttributeError:
            return False

# --- UI COMPONENTS & DIALOGS ---

class BaseDialog(tk.Toplevel):
    def __init__(self, parent, title="Dialog"):
        super().__init__(parent)
        self.title(title)
        self.result = None
        self.transient(parent)
        self.grab_set()
        self.body = ttk.Frame(self, padding=20)
        self.body.pack(fill="both", expand=True)
        self._setup_ui()
        btn_frame = ttk.Frame(self, padding=(0,0,20,20))
        btn_frame.pack(side="bottom", anchor="e")
        ttk.Button(btn_frame, text="Cancel", command=self.destroy).pack(side="right", padx=5)
        ttk.Button(btn_frame, text="Save", command=self._on_save).pack(side="right")

    def _setup_ui(self): pass
    def _on_save(self): pass

class EmployeeDialog(BaseDialog):
    def __init__(self, parent, emp=None):
        self.emp = emp
        super().__init__(parent, "Employee Details")

    def _setup_ui(self):
        self.vars = {}
        fields = [
            ("emp_id", "Employee ID"), ("name", "Name"), ("role", "Role"),
            ("department", "Department"), ("contract_type", "Contract Type"),
            ("contracted_weekly_hrs", "Weekly Hours"), ("shift_eligibility", "Shift Eligibility"),
            ("is_hazardous", "Haz (0/1)"), ("is_industrial_rotating", "Ind Rot (0/1)"),
            ("hour_exempt", "Exempt (0/1)"), ("fixed_rest_day", "Rest Day (1-7)"),
            ("phone", "Phone"), ("hire_date", "Hire Date (YYYY-MM-DD)"), ("notes", "Notes")
        ]
        for idx, (attr, label) in enumerate(fields):
            ttk.Label(self.body, text=label).grid(row=idx, column=0, sticky="w", pady=2)
            var = tk.StringVar()
            if self.emp:
                val = getattr(self.emp, attr)
                var.set(str(val) if val is not None else "")
            self.vars[attr] = var
            ttk.Entry(self.body, textvariable=var, width=40).grid(row=idx, column=1, padx=5)

    def _on_save(self):
        try:
            d_str = self.vars["hire_date"].get()
            hd = datetime.strptime(d_str, "%Y-%m-%d").date() if d_str else None
            self.result = Employee(
                self.vars["emp_id"].get(), self.vars["name"].get(), 
                self.vars["role"].get(), self.vars["department"].get(),
                self.vars["contract_type"].get(), float(self.vars["contracted_weekly_hrs"].get()),
                self.vars["shift_eligibility"].get(), int(self.vars["is_hazardous"].get()),
                int(self.vars["is_industrial_rotating"].get()), int(self.vars["hour_exempt"].get()),
                int(self.vars["fixed_rest_day"].get()), self.vars["phone"].get(), hd, self.vars["notes"].get()
            )
            self.destroy()
        except Exception as e:
            messagebox.showerror("Validation Error", str(e))

class ShiftDialog(BaseDialog):
    def __init__(self, parent, shift=None):
        self.shift = shift
        super().__init__(parent, "Shift Configuration")

    def _setup_ui(self):
        self.vars = {}
        fields = [
            ("code", "Code"), ("name", "Name"), ("start_time", "Start (HH:MM)"),
            ("end_time", "End (HH:MM)"), ("duration_hrs", "Duration Hrs"),
            ("break_min", "Break (min)"), ("is_industrial", "Is Industrial (0/1)"),
            ("is_hazardous", "Is Hazardous (0/1)"), ("is_work", "Is Work (0/1)"),
            ("description", "Description")
        ]
        for idx, (attr, label) in enumerate(fields):
            ttk.Label(self.body, text=label).grid(row=idx, column=0, sticky="w", pady=2)
            var = tk.StringVar()
            if self.shift:
                val = getattr(self.shift, attr)
                var.set(str(val) if val is not None else "")
            self.vars[attr] = var
            ttk.Entry(self.body, textvariable=var, width=40).grid(row=idx, column=1, padx=5)

    def _on_save(self):
        try:
            s_t = datetime.strptime(self.vars["start_time"].get(), "%H:%M").time()
            e_t = datetime.strptime(self.vars["end_time"].get(), "%H:%M").time()
            self.result = Shift(
                self.vars["code"].get(), self.vars["name"].get(), s_t, e_t,
                float(self.vars["duration_hrs"].get()), int(self.vars["break_min"].get()),
                int(self.vars["is_industrial"].get()), int(self.vars["is_hazardous"].get()),
                int(self.vars["is_work"].get()), self.vars["description"].get()
            )
            self.destroy()
        except Exception as e:
            messagebox.showerror("Validation Error", str(e))

# --- MAIN TABS ---

class DashboardTab(ttk.Frame):
    def __init__(self, parent, store):
        super().__init__(parent)
        self.store = store
        self._setup_ui()

    def _setup_ui(self):
        self.stat_frame = ttk.Frame(self, padding=10)
        self.stat_frame.pack(fill="x", side="top")
        self.kpi_labels = {}
        kpis = ["Total Employees", "Avg Monthly Hours", "Total Violations", "Peak-day Coverage Avg", "Critical Site Count"]
        for i, kpi in enumerate(kpis):
            f = ttk.LabelFrame(self.stat_frame, text=kpi, padding=10)
            f.grid(row=0, column=i, padx=5, sticky="nsew")
            lbl = ttk.Label(f, text="0", font=("TkDefaultFont", 16, "bold"))
            lbl.pack()
            self.kpi_labels[kpi] = lbl

        self.tree = ttk.Treeview(self, columns=("EmpID", "Name", "TotalHrs", "Violations"), show="headings")
        for c in ("EmpID", "Name", "TotalHrs", "Violations"):
            self.tree.heading(c, text=c)
            self.tree.column(c, width=120)
        self.tree.pack(fill="both", expand=True, padx=10, pady=10)
        self.tree.tag_configure("violation", background="#ffcccc")

    def refresh(self):
        for item in self.tree.get_children(): self.tree.delete(item)
        violations = ComplianceEngine.check_compliance(self.store)
        v_counts = {}
        for v in violations: v_counts[v.emp_id] = v_counts.get(v.emp_id, 0) + 1
        
        total_hrs_total = 0
        shifts_by_code = {s.code: s for s in self.store.shifts}
        
        for emp in self.store.employees:
            sched = self.store.schedule.get(emp.emp_id, {})
            emp_hrs = sum(shifts_by_code[c].duration_hrs for c in sched.values() if c in shifts_by_code)
            total_hrs_total += emp_hrs
            v_count = v_counts.get(emp.emp_id, 0)
            tag = ("violation",) if v_count > 0 else ()
            self.tree.insert("", "end", values=(emp.emp_id, emp.name, f"{emp_hrs:.1f}", v_count), tags=tag)

        self.kpi_labels["Total Employees"].config(text=str(len(self.store.employees)))
        avg_hrs = total_hrs_total / len(self.store.employees) if self.store.employees else 0
        self.kpi_labels["Avg Monthly Hours"].config(text=f"{avg_hrs:.1f}")
        self.kpi_labels["Total Violations"].config(text=str(len(violations)))
        
        # Peak day calc (Thu/Fri/Sat = 5,6,7)
        peak_cov = []
        for d in range(1, self.store.config.days_in_month + 1):
            try:
                dt = date(self.store.config.year, self.store.config.month, d)
                if dt.isoweekday() in [4, 5, 6]: # Sunday=1 map
                    # Iraqi convention start=Sun(1), Fri=6, Sat=7. 
                    # Python isoweekday: Mon=1, Thu=4, Fri=5, Sat=6, Sun=7.
                    # Iraqi Thu/Fri/Sat are 5,6,7 which map to Python 4,5,6?
                    count = sum(1 for s in self.store.schedule.values() if s.get(d) and shifts_by_code.get(s.get(d), Shift("", "", time(0,0), time(0,0), 0)).is_work)
                    peak_cov.append(count)
            except: continue
        avg_peak = sum(peak_cov)/len(peak_cov) if peak_cov else 0
        self.kpi_labels["Peak-day Coverage Avg"].config(text=f"{avg_peak:.1f}")

class RosterTab(ttk.Frame):
    def __init__(self, parent, store):
        super().__init__(parent)
        self.store = store
        self._setup_ui()

    def _setup_ui(self):
        btn_frame = ttk.Frame(self, padding=5)
        btn_frame.pack(side="top", fill="x")
        ttk.Button(btn_frame, text="Add Employee", command=self._add).pack(side="left", padx=5)
        ttk.Button(btn_frame, text="Edit Selected", command=self._edit).pack(side="left", padx=5)
        ttk.Button(btn_frame, text="Delete Selected", command=self._delete).pack(side="left", padx=5)

        cols = ("ID", "Name", "Role", "Dept", "Contract", "Hrs", "Haz")
        self.tree = ttk.Treeview(self, columns=cols, show="headings")
        for c in cols: 
            self.tree.heading(c, text=c)
            self.tree.column(c, width=100)
        self.tree.pack(fill="both", expand=True, padx=10, pady=5)

    def refresh(self):
        for item in self.tree.get_children(): self.tree.delete(item)
        for emp in self.store.employees:
            self.tree.insert("", "end", values=(emp.emp_id, emp.name, emp.role, emp.department, 
                                               emp.contract_type, emp.contracted_weekly_hrs, emp.is_hazardous))

    def _add(self):
        d = EmployeeDialog(self.winfo_toplevel())
        self.wait_window(d)
        if d.result:
            self.store.employees.append(d.result)
            self.store.dirty = True
            self.refresh()

    def _edit(self):
        sel = self.tree.selection()
        if not sel: return
        emp_id = self.tree.item(sel[0])["values"][0]
        emp = next((e for e in self.store.employees if e.emp_id == str(emp_id)), None)
        if emp:
            d = EmployeeDialog(self.winfo_toplevel(), emp)
            self.wait_window(d)
            if d.result:
                idx = self.store.employees.index(emp)
                self.store.employees[idx] = d.result
                self.store.dirty = True
                self.refresh()

    def _delete(self):
        sel = self.tree.selection()
        if not sel: return
        if messagebox.askyesno("Confirm", "Delete selected employee?"):
            emp_id = self.tree.item(sel[0])["values"][0]
            self.store.employees = [e for e in self.store.employees if e.emp_id != str(emp_id)]
            self.store.dirty = True
            self.refresh()

class ScheduleGridTab(ttk.Frame):
    def __init__(self, parent, store):
        super().__init__(parent)
        self.store = store
        self._setup_ui()

    def _setup_ui(self):
        cols = ("EmpID", "Name") + tuple(range(1, 32))
        self.tree = ttk.Treeview(self, columns=cols, show="headings")
        self.tree.heading("EmpID", text="ID")
        self.tree.column("EmpID", width=60, stretch=False)
        self.tree.heading("Name", text="Name")
        self.tree.column("Name", width=120, stretch=False)
        for i in range(1, 32):
            self.tree.heading(i, text=str(i))
            self.tree.column(i, width=35, anchor="center")
        
        self.tree.pack(fill="both", expand=True, padx=10, pady=10)
        self.tree.bind("<Double-1>", self._on_double_click)

        # Tags for shift colors
        self.tree.tag_configure("FS", foreground="blue", font=("TkDefaultFont", 10, "bold"))
        self.tree.tag_configure("HS", foreground="green", font=("TkDefaultFont", 10, "bold"))
        self.tree.tag_configure("MX", foreground="orange", font=("TkDefaultFont", 10, "bold"))
        self.tree.tag_configure("OFF", foreground="grey")
        self.tree.tag_configure("PH", foreground="red")
        self.tree.tag_configure("AL", foreground="purple")
        self.tree.tag_configure("SL", foreground="#CCCC00")

    def refresh(self):
        for item in self.tree.get_children(): self.tree.delete(item)
        for emp in self.store.employees:
            sched = self.store.schedule.get(emp.emp_id, {})
            row_vals = [emp.emp_id, emp.name]
            for d in range(1, 32):
                row_vals.append(sched.get(d, ""))
            # Just tag by the first shift code for simplicity, or complex tagging
            self.tree.insert("", "end", values=row_vals)

    def _on_double_click(self, event):
        item = self.tree.identify_item(event.y)
        column = self.tree.identify_column(event.x)
        if not item or not column: return
        col_idx = int(column.replace("#", ""))
        if col_idx < 3: return # EmpID or Name
        
        day = col_idx - 2
        emp_id = str(self.tree.item(item)["values"][0])
        
        pop = tk.Toplevel(self)
        pop.title("Pick Shift")
        codes = [s.code for s in self.store.shifts] + ["OFF", "AL", "SL", "PH", ""]
        combo = ttk.Combobox(pop, values=codes)
        combo.pack(padx=20, pady=20)
        combo.set(self.store.schedule.get(emp_id, {}).get(day, ""))
        
        def save():
            new_val = combo.get()
            if emp_id not in self.store.schedule: self.store.schedule[emp_id] = {}
            if new_val: self.store.schedule[emp_id][day] = new_val
            elif day in self.store.schedule[emp_id]: del self.store.schedule[emp_id][day]
            self.store.dirty = True
            self.refresh()
            pop.destroy()

        ttk.Button(pop, text="OK", command=save).pack(pady=5)

class ShiftsTab(ttk.Frame):
    def __init__(self, parent, store):
        super().__init__(parent)
        self.store = store
        self._setup_ui()

    def _setup_ui(self):
        btn_frame = ttk.Frame(self, padding=5)
        btn_frame.pack(side="top", fill="x")
        ttk.Button(btn_frame, text="Add Shift", command=self._add).pack(side="left", padx=5)
        ttk.Button(btn_frame, text="Edit Shift", command=self._edit).pack(side="left", padx=5)
        
        cols = ("Code", "Name", "Start", "End", "Hrs", "Break", "Ind", "Haz", "Work")
        self.tree = ttk.Treeview(self, columns=cols, show="headings")
        for c in cols:
            self.tree.heading(c, text=c)
            self.tree.column(c, width=70)
        self.tree.pack(fill="both", expand=True, padx=10, pady=5)

    def refresh(self):
        for item in self.tree.get_children(): self.tree.delete(item)
        for s in self.store.shifts:
            self.tree.insert("", "end", values=(s.code, s.name, s.start_time, s.end_time, 
                                               s.duration_hrs, s.break_min, s.is_industrial, 
                                               s.is_hazardous, s.is_work))

    def _add(self):
        d = ShiftDialog(self.winfo_toplevel())
        self.wait_window(d)
        if d.result:
            self.store.shifts.append(d.result)
            self.store.dirty = True
            self.refresh()

    def _edit(self):
        sel = self.tree.selection()
        if not sel: return
        code = self.tree.item(sel[0])["values"][0]
        s = next((x for x in self.store.shifts if x.code == str(code)), None)
        if s:
            d = ShiftDialog(self.winfo_toplevel(), s)
            self.wait_window(d)
            if d.result:
                idx = self.store.shifts.index(s)
                self.store.shifts[idx] = d.result
                self.store.dirty = True
                self.refresh()

class CoverageTab(ttk.Frame):
    def __init__(self, parent, store):
        super().__init__(parent)
        self.store = store
        self.canvas = tk.Canvas(self, bg="white")
        self.canvas.pack(fill="both", expand=True)

    def refresh(self):
        self.canvas.delete("all")
        codes = sorted([s.code for s in self.store.shifts if s.is_work])
        if not codes: return
        
        cell_w, cell_h = 30, 30
        margin_x, margin_y = 60, 40
        
        # Draw headers
        for d in range(1, 32):
            self.canvas.create_text(margin_x + d*cell_w + 15, margin_y - 15, text=str(d))
        for i, code in enumerate(codes):
            self.canvas.create_text(margin_x - 30, margin_y + i*cell_h + 15, text=code)

        # Draw Heatmap
        for i, code in enumerate(codes):
            for d in range(1, 32):
                count = sum(1 for s in self.store.schedule.values() if s.get(d) == code)
                color = "white"
                if count > 0:
                    intensity = min(255, count * 25)
                    color = f"#{255-intensity:02x}ff{255-intensity:02x}"
                if count < self.store.config.coverage_min:
                    color = "#ffcccc"
                
                x, y = margin_x + d*cell_w, margin_y + i*cell_h
                self.canvas.create_rectangle(x, y, x+cell_w, y+cell_h, fill=color, outline="#ddd")
                self.canvas.create_text(x+15, y+15, text=str(count))

class SettingsTab(ttk.Frame):
    def __init__(self, parent, store):
        super().__init__(parent)
        self.store = store
        self._setup_ui()

    def _setup_ui(self):
        self.vars = {}
        row = 0
        fields = [
            ("company", "Company Name"), ("year", "Year"), ("month", "Month (1-12)"),
            ("coverage_min", "Min Coverage Per Shift"), ("max_consec_work_days", "Max Consec Work Days"),
            ("continuous_shifts_mode", "Art. 71 Mode (ON/OFF)")
        ]
        container = ttk.Frame(self, padding=20)
        container.pack(fill="both")
        for attr, label in fields:
            ttk.Label(container, text=label).grid(row=row, column=0, sticky="w", pady=5)
            var = tk.StringVar()
            self.vars[attr] = var
            ttk.Entry(container, textvariable=var, width=30).grid(row=row, column=1, padx=10)
            row += 1
        ttk.Button(container, text="Apply & Save Config", command=self._save).grid(row=row, column=1, sticky="e", pady=20)

    def refresh(self):
        for attr, var in self.vars.items():
            var.set(str(getattr(self.store.config, attr)))

    def _save(self):
        try:
            self.store.config.company = self.vars["company"].get()
            self.store.config.year = int(self.vars["year"].get())
            self.store.config.month = int(self.vars["month"].get())
            self.store.config.coverage_min = int(self.vars["coverage_min"].get())
            self.store.config.max_consec_work_days = int(self.vars["max_consec_work_days"].get())
            self.store.config.continuous_shifts_mode = self.vars["continuous_shifts_mode"].get()
            self.store.dirty = True
            messagebox.showinfo("Success", "Settings applied.")
        except Exception as e:
            messagebox.showerror("Error", str(e))

class ReportsTab(ttk.Frame):
    def __init__(self, parent, store):
        super().__init__(parent)
        self.store = store
        ttk.Label(self, text="Export Monthly Reports", font=("TkDefaultFont", 12, "bold")).pack(pady=20)
        ttk.Button(self, text="Export Schedule CSV", command=self._sched).pack(pady=10)
        ttk.Button(self, text="Export Violations CSV", command=self._viol).pack(pady=10)
        ttk.Button(self, text="Export Coverage CSV", command=self._cov).pack(pady=10)

    def _sched(self):
        f = filedialog.asksaveasfilename(defaultextension=".csv")
        if not f: return
        with open(f, "w", newline="") as csvf:
            writer = csv.writer(csvf)
            writer.writerow(["EmpID", "Name"] + list(range(1, 32)))
            for emp in self.store.employees:
                s = self.store.schedule.get(emp.emp_id, {})
                writer.writerow([emp.emp_id, emp.name] + [s.get(d, "") for d in range(1, 32)])

    def _viol(self):
        f = filedialog.asksaveasfilename(defaultextension=".csv")
        if not f: return
        viols = ComplianceEngine.check_compliance(self.store)
        with open(f, "w", newline="") as csvf:
            writer = csv.writer(csvf)
            writer.writerow(["EmpID", "Day", "Rule", "Article", "Message"])
            for v in viols:
                writer.writerow([v.emp_id, v.day, v.rule, v.article, v.message])

    def _cov(self):
        f = filedialog.asksaveasfilename(defaultextension=".csv")
        if not f: return
        codes = [s.code for s in self.store.shifts if s.is_work]
        with open(f, "w", newline="") as csvf:
            writer = csv.writer(csvf)
            writer.writerow(["ShiftCode"] + list(range(1, 32)))
            for c in codes:
                row = [c]
                for d in range(1, 32):
                    row.append(sum(1 for s in self.store.schedule.values() if s.get(d) == c))
                writer.writerow(row)

# --- MAIN APPLICATION ---

class SchedulerApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Iraqi Labor Law Workforce Scheduler — AL-Rafidain Group")
        self.geometry("1200x800")
        tkfont.nametofont("TkDefaultFont").configure(size=10)
        
        self.store = DataStore()
        self._setup_menu()
        self._setup_ui()
        
        self.status = tk.StringVar(value="Ready")
        sb = ttk.Label(self, textvariable=self.status, relief="sunken", anchor="w")
        sb.pack(side="bottom", fill="x")

        # Auto-load template if exists
        if os.path.exists("Scheduler_Template.xlsx"):
            self._load_file("Scheduler_Template.xlsx")

    def _setup_menu(self):
        m = tk.Menu(self)
        fm = tk.Menu(m, tearoff=0)
        fm.add_command(label="Open Excel", command=self._open)
        fm.add_command(label="Save", command=self._save)
        fm.add_command(label="Save As...", command=self._save_as)
        fm.add_separator()
        fm.add_command(label="Exit", command=self.quit)
        m.add_cascade(label="File", menu=fm)
        self.config(menu=m)

    def _setup_ui(self):
        self.nb = ttk.Notebook(self)
        self.nb.pack(fill="both", expand=True)
        
        self.tabs = {
            "Dashboard": DashboardTab(self.nb, self.store),
            "Roster": RosterTab(self.nb, self.store),
            "Schedule": ScheduleGridTab(self.nb, self.store),
            "Shifts": ShiftsTab(self.nb, self.store),
            "Coverage": CoverageTab(self.nb, self.store),
            "Reports": ReportsTab(self.nb, self.store),
            "Settings": SettingsTab(self.nb, self.store),
            "Holidays": ttk.Frame(self.nb) # Placeholder
        }
        
        for name, tab in self.tabs.items():
            self.nb.add(tab, text=name)
        
        self.nb.bind("<<NotebookTabChanged>>", lambda e: self._refresh_current())

    def _refresh_current(self):
        idx = self.nb.index("current")
        tab_name = self.nb.tab(idx, "text")
        tab = self.tabs.get(tab_name)
        if tab and hasattr(tab, "refresh"):
            tab.refresh()
        mod = "*" if self.store.dirty else ""
        self.status.set(f"File: {self.store.filename or 'None'} {mod}")

    def _load_file(self, path):
        try:
            self.store.load(path)
            self._refresh_current()
            self.status.set(f"Loaded {path}")
        except Exception as e:
            messagebox.showerror("Load Error", str(e))

    def _open(self):
        f = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if f: self._load_file(f)

    def _save(self):
        try:
            if not self.store.filename:
                return self._save_as()
            self.store.save()
            self.status.set("Saved.")
            self._refresh_current()
        except Exception as e:
            messagebox.showerror("Save Error", str(e))

    def _save_as(self):
        f = filedialog.asksaveasfilename(defaultextension=".xlsx")
        if f:
            self.store.save(f)
            self.status.set(f"Saved as {f}")
            self._refresh_current()

if __name__ == "__main__":
    app = SchedulerApp()
    app.mainloop()
